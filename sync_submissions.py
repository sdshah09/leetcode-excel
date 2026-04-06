#!/usr/bin/env python3
"""
Daily LeetCode → Excel sync for GitHub Actions.
Uses public GraphQL API - no authentication required.

Usage:
  LEETCODE_USERNAME=sdshah0616 python sync_submissions.py

Environment variables:
  LEETCODE_USERNAME  — your LeetCode username (required)
  EXCEL_PATH         — path to workbook (default: data/leetcode_log.xlsx)

Note: LeetCode returns max 20 recent accepted submissions.
"""

import os
import sys
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

GRAPHQL_URL = "https://leetcode.com/graphql/"

RECENT_AC_QUERY = """
query recentAcSubmissions($username: String!, $limit: Int!) {
  recentAcSubmissionList(username: $username, limit: $limit) {
    id
    title
    titleSlug
    timestamp
  }
}
"""


def fetch_accepted_submissions(username: str, limit: int = 20) -> list[dict]:
    """Fetch recent accepted submissions using public API."""
    session = requests.Session()

    # Get CSRF token first
    session.get(
        f"https://leetcode.com/u/{username}/",
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"},
        timeout=30
    )
    csrf = session.cookies.get("csrftoken", "")

    response = session.post(
        GRAPHQL_URL,
        json={
            "operationName": "recentAcSubmissions",
            "query": RECENT_AC_QUERY,
            "variables": {"username": username, "limit": limit}
        },
        headers={
            "Content-Type": "application/json",
            "Origin": "https://leetcode.com",
            "Referer": f"https://leetcode.com/u/{username}/",
            "x-csrftoken": csrf,
        },
        timeout=30
    )
    response.raise_for_status()

    data = response.json()
    if "errors" in data:
        raise RuntimeError(f"GraphQL error: {data['errors']}")

    return data.get("data", {}).get("recentAcSubmissionList") or []


def group_by_day(submissions: list[dict]) -> dict[date, set[str]]:
    """Group submissions by date."""
    by_day: dict[date, set[str]] = defaultdict(set)

    for sub in submissions:
        title = sub.get("title", "").strip()
        if not title:
            continue
        ts = int(sub["timestamp"])
        d = datetime.fromtimestamp(ts, tz=timezone.utc).date()
        by_day[d].add(title)

    return dict(by_day)


def load_existing(path: Path) -> dict[date, set[str]]:
    """Load existing entries from Excel file."""
    if not path.exists():
        return {}

    wb = load_workbook(path)
    ws = wb["Log"] if "Log" in wb.sheetnames else wb.active
    existing: dict[date, set[str]] = {}

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[0] is None:
            continue

        raw_date = row[0]
        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            d = date.fromisoformat(str(raw_date)[:10])

        names_cell = row[3]
        if names_cell and str(names_cell).strip():
            names = {p.strip() for p in str(names_cell).split("\n") if p.strip()}
        else:
            names = set()
        existing[d] = names

    return existing


def save_workbook(path: Path, data: dict[date, set[str]]) -> None:
    """Save data to Excel workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["Date", "Day", "Problems Solved", "Problem Names"])

    for d in sorted(data.keys()):
        titles = sorted(data[d])
        ws.append([d, d.strftime("%A"), len(titles), "\n".join(titles)])

    # Set column widths
    ws.column_dimensions["A"].width = 12  # Date
    ws.column_dimensions["B"].width = 12  # Day
    ws.column_dimensions["C"].width = 16  # Problems Solved
    ws.column_dimensions["D"].width = 50  # Problem Names

    wb.save(path)


def main() -> int:
    username = os.environ.get("LEETCODE_USERNAME", "").strip()
    if not username:
        print("Error: LEETCODE_USERNAME is required", file=sys.stderr)
        print("Usage: LEETCODE_USERNAME=your_username python sync_submissions.py")
        return 1

    excel_path = Path(os.environ.get("EXCEL_PATH", "data/leetcode_log.xlsx"))

    print(f"Fetching submissions for {username}...")
    submissions = fetch_accepted_submissions(username, limit=20)

    if not submissions:
        print("No accepted submissions found.")
        return 0

    print(f"Found {len(submissions)} recent accepted submissions (max 20)")

    # Group by day and merge with existing
    from_api = group_by_day(submissions)
    existing = load_existing(excel_path)

    merged = {d: set(names) for d, names in existing.items()}
    for d, titles in from_api.items():
        merged.setdefault(d, set()).update(titles)

    if existing == merged:
        print(f"No new submissions to add. Excel has {len(merged)} days.")
        return 0

    save_workbook(excel_path, merged)
    new_days = len(merged) - len(existing)
    print(f"Updated {excel_path} - {len(merged)} total days ({new_days} new)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
